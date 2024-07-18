import openpyxl
import copy
import glob
import os
from openpyxl.utils import get_column_letter

work_dir = r"E:\\Python\\merge_sheet"
file_name = '*.xlsx'

def merge_excel(work_dir, file_name):
    file_pattern = os.path.join(work_dir, file_name)
    # 创建一个新的工作表
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.create_sheet('Merged',0)
    # 遍历所有excel文件的所有sheet, 存为list
    wb_list = []
    sheet_list = []
    for f in glob.glob(file_pattern):
        wb1 = openpyxl.load_workbook(f)
        wb_list.append(wb1)
        for sheet in wb1:
            sheet_list.append(sheet)

    cur_row = 1
    for i,sheet in enumerate(sheet_list):
        for row in range(sheet.max_row):     # 遍历行
            # 设置行高。 新sheet页行高=旧sheet页行高。
            new_sheet.row_dimensions[cur_row].height = sheet.row_dimensions[row+1].height
            for col in range(sheet.max_column):   # 遍历列
                # 设置列宽。 新sheet页列宽=旧sheet页列宽。
                new_sheet.column_dimensions[get_column_letter(col+1)].width = sheet.column_dimensions[get_column_letter(col+1)].width
                # 设置单元格的值
                new_sheet.cell(row=cur_row,column=col+1,value=sheet.cell(row+1,col+1).value)
                # 拷贝格式
                if sheet.cell(row+1,col+1).has_style:
                    new_sheet.cell(row=cur_row,column=col+1).font = copy.copy(sheet.cell(row+1,col+1).font)
                    new_sheet.cell(row=cur_row,column=col+1).border = copy.copy(sheet.cell(row+1,col+1).border)
                    new_sheet.cell(row=cur_row,column=col+1).fill = copy.copy(sheet.cell(row+1,col+1).fill)
                    new_sheet.cell(row=cur_row,column=col+1).number_format = copy.copy(sheet.cell(row+1,col+1).number_format)
                    new_sheet.cell(row=cur_row,column=col+1).protection = copy.copy(sheet.cell(row+1,col+1).protection)
                    new_sheet.cell(row=cur_row,column=col+1).alignment = copy.copy(sheet.cell(row+1,col+1).alignment)
            cur_row += 1

    # 保存新的Excel文件
    new_wb.save(os.path.join(work_dir,'merged.xlsx'))
    print("save excel to: " + os.path.join(work_dir,'merged.xlsx'))


merge_excel(work_dir, file_name)