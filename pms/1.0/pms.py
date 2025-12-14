import sys
import re
import os
import glob
import openpyxl
import shutil
import getpass
from dataclasses import dataclass,field,fields
from pathlib import Path
from datetime import datetime
from typing import Optional

#============================================================
# Define data class
@dataclass
class Path:
    top:str="None"
    top_path:str="None"
    rtl_path:str="None"
    rtl_list:str="None"
    tb_path:str="None"
    cons_path:str="None"
    ip_path:str="None"
    mem_path:str="None"
    param:str="None"
    port_define:str="None"
    lib_path:str="None"
    sub_rtl_list:str="None"
    sub_path_list:str="None"
    sub_path_list_raw:str="None"
    mem_list:str="None"
@dataclass
class Range:
    min:str="None"
    max:str="None"
@dataclass
class Library:
    post_sim:str="None"
    target:str="None"
    link:str="None"
    symbol:str="None"
    synthetic:str="None"
@dataclass
class DrivingCell:
    name:str="None"
    input:str="None"
    output:str="None"
    lib:str="None"
@dataclass
class ClockTree:
    source_latency:Range=field(default_factory=Range)
    network_latency:Range=field(default_factory=Range)
    trans:Range=field(default_factory=Range)
    skew:str="None"
    noise:str="None"
@dataclass
class Scale:
    sm:ClockTree=field(default_factory=ClockTree)
    md:ClockTree=field(default_factory=ClockTree)
    lg:ClockTree=field(default_factory=ClockTree)
    raw:ClockTree=field(default_factory=ClockTree)
@dataclass
class Config:
    path_root:str="None"
    path_sub:str="None"
    library:Library=field(default_factory=Library)
    signal_driving_cell:DrivingCell=field(default_factory=DrivingCell)
    clock_driving_cell:DrivingCell=field(default_factory=DrivingCell)
    signal_trans:Range=field(default_factory=Range)
    output_load:Range=field(default_factory=Range)
    output_trans:Range=field(default_factory=Range)
    output_delay:Range=field(default_factory=Range)
    input_delay:Range=field(default_factory=Range)
    input_trans:Range=field(default_factory=Range)
    wire_load_model:str="None"
    opera_condition:str="None"
    setup_margin:str="None"
    hold_margin:str="None"
    ct:Scale=field(default_factory=Scale)

def replace_in_file(file_path,target,replacement):
    if replacement is None:
        replacement = ""
    with open(file_path,'r+',encoding='utf-8') as f:
        content = f.read()
        f.seek(0)
        f.write(content.replace(target,replacement))
        f.truncate()
def pms_msg(msg):
    print(f"[PMS] {msg}")
def pms_info(msg):
    pms_msg(f"   INFO: {msg}")
def pms_warning(msg):
    pms_msg(f"WARNING: {msg}")
def pms_error(msg):
    pms_msg(f"  ERROR: {msg}")
def pms_fatal(msg):
    print(f"[PMS]   FATAL: {msg}", file=sys.stderr)
    sys.exit(1)
def print_startup():
    time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    username=getpass.getuser()
    version = os.path.basename(os.path.dirname(os.path.abspath(__file__)))
    pms_msg("#"+"="*60)
    pms_msg(f"# PMS Version {version}")
    pms_msg("#"+"-"*60)
    pms_msg(f"# Date:        {time}")
    pms_msg(f"# Owner:       {username}")
    pms_msg("#"+"="*60)
    print()
def standardize_list(list):
    if not list:
        return ""
    elif list is None:
        return ""
    return list.replace(","," ").replace(";"," ").replace("\n"," ").replace("\r"," ")
def check_dir(parent_dir,sub_dir=None):
    print("")

    if os.path.exists(parent_dir):
        if "design/tb" in parent_dir:
            pms_warning("Testbench directory existed, will not change.")
            return 0
        else:
            user_input = input(f"[PMS] WARNING: DIR {parent_dir} existed, do you want to clean it and continue?(y/n):").strip().lower()
            if user_input == "n":
                pms_info(f"No change for {parent_dir}\n")
                return 0
            elif user_input == "y":
                shutil.rmtree(parent_dir)
            else:
                pms_fatal("Input error.")

    
    os.makedirs(parent_dir)
    if sub_dir:
        for each_sub_dir in sub_dir:
            os.makedirs(os.path.join(parent_dir,each_sub_dir))
    pms_info(f"Create {parent_dir} sucessfully\n")
    return 1

def get_env_var(var,bak):
    if var in os.environ:
        return os.environ.get(var)
    else:
        return bak
def get_rtl_file(path):
    rtl_list = []
    rtl_pattern = ["*.v", "*.sv", "*.vhd", "*.vhdl"]
    for pattern in rtl_pattern:
        rtl_list.extend(glob.glob(os.path.join(path, pattern)))
    if not rtl_list:
        pms_fatal(f"Not find RTL files in {path} with pattern {rtl_pattern}")
    rtl_list = " ".join(os.path.basename(f) for f in rtl_list)
    return rtl_list
def get_port_define(rtl_file):
    with open(rtl_file,'r') as f:
        content = f.read()
    content = re.sub(r'\/\*.*?\*\/','',content,flags=re.DOTALL)
    content = re.sub(r'\/\/.*','',content)
    match = re.search(r'module\s+\w+\s*(?:#\s*\(.*?\)\s*)?\((.*?)\);\s*',content,re.DOTALL)
    if not match:
        pms_fatal("Can not find port define")
    raw_port_define = match.group(1)
    port2logic = re.sub(r'\b(input\s+wire|output\s+reg|output\s+logic|output\s+wire|input|output|inout|reg|wire)\b','logic',raw_port_define)
    port2logic = re.sub(r'\blogic\s+logic\b','logic',port2logic)
    port2logic = re.sub(r'(?m)^[ \t\f\r\v]*(?:\r?\n|$)','',port2logic)
    port2logic = re.sub(r'(?m)[ \t\f\r\v]+$','',port2logic)
    port2logic = re.sub(r'(?m)[,;]+$','',port2logic)
    port2logic = re.sub(r'(?m)(\S)$',r'\1;',port2logic)
    return port2logic
def get_top(sheet):
    top = sheet['B1'].value
    if top is None or str(top).strip() == "":
        pms_fatal("Not define top module.")
    return top
def get_top_path(top,sheet):
    top_path = get_env_var(top,sheet['B2'].value)
    top_path = re.sub(r"\$\{top\}",top,top_path,flags=re.IGNORECASE)
    if top_path is None:
        pms_fatal("Error: Not define design path")
    return top_path
def get_top_rtl_path(top_path,sheet):
    top_rtl_path = sheet['B3'].value
    top_rtl_path = re.sub(r"\$\{top_path\}",top_path,top_rtl_path,flags=re.IGNORECASE)
    return top_rtl_path
def get_sub_rtl(sheet):
    sub_path_list = ""
    sub_rtl_list = ""
    sub_rtl_list_raw = ""
    row_index   = 11
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        sub_path = get_env_var(row[0],row[1])
        sub_sheet = openpyxl.load_workbook(os.path.join(sub_path,"pms",row[0]+".xlsx"))['path']
        sub_top_path = get_top_rtl_path(get_top_path(get_top(sub_sheet),sub_sheet),sub_sheet)
        sub_path_list += f" {sub_top_path}"
        sub_rtl_list_current = get_rtl_file(sub_top_path)
        sub_rtl_list += " " + sub_rtl_list_current
        sub_rtl_list_raw += ' '+' '.join([sub_top_path+'/'+rtl_name for rtl_name in sub_rtl_list_current.split()])
        sub_sub_path_list,sub_sub_rtl_list,sub_sub_rtl_list_raw = get_sub_rtl(sub_sheet)
        sub_path_list += sub_sub_path_list
        sub_rtl_list += sub_sub_rtl_list
        sub_rtl_list_raw += sub_sub_rtl_list_raw
        row_index += 1
    return sub_path_list,sub_rtl_list,sub_rtl_list_raw
def get_mem_list(sheet):
    mem_list = []
    row_index = 2 
    while True:
        cell_value = sheet.cell(row=row_index, column=10).value
        
        if cell_value is None or str(cell_value).strip() == "":
            break
            
        mem_list.append(cell_value)
        row_index += 1 

    return mem_list
def find_single_excel(path):
    if not os.path.isdir(path):
        pms_fatal(f"路径不存在或不是目录: {path}")

    excel_files = (
        glob.glob(os.path.join(path, "*.xls")) +
        glob.glob(os.path.join(path, "*.xlsx"))
    )
    if len(excel_files) != 1:
        pms_fatal(
            f"在路径 {path} 下找到 {len(excel_files)} 个 Excel 文件，"
        )
    return excel_files[0]
def get_all_mem(mem_path,sub_path_list
,sheet):
    mem_list = get_mem_list(sheet)
    path_list = sub_path_list.split()
    if not path_list:
        return []
    for each_path in path_list:
        full_path = os.path.join(each_path, "../../pms")
        print(full_path)
        sub_sheet = openpyxl.load_workbook(find_single_excel(full_path))['path']
        mem_list.extend(get_mem_list(sub_sheet))
    return mem_list

def parse_path(sheet):
    pms_msg("#"+"-"*60)
    pms_msg("# Fetch design and path information")
    pms_msg("#"+"-"*60)
    path=Path()
    path.top = get_top(sheet)
    pms_info(f"Design top module: {path.top}")
    path.top_path = get_top_path(path.top,sheet)
    pms_info(f"Top dir: {path.top_path}")
    path.rtl_path = get_top_rtl_path(path.top_path,sheet)
    path.rtl_list = get_rtl_file(path.rtl_path)

    top_pattern = path.top+"."
    top_match = [item for item in path.rtl_list.split() if top_pattern in item]
    if len(top_match) == 1:
        top_file = top_match[0]
    else:
        pms_error(f"Top-level file: {top_match}")
        pms_fatal("The top-level file is not unique or not found.")

    param_lines = []
    found_param = False
    with open(path.rtl_path+"/"+top_file,'r',encoding='utf-8') as file:
        for line in file:
            stripped = line.strip()
            if stripped.startswith("//"):
                continue
            stripped = stripped.split('//')[0].rstrip()
            if not found_param:
                if "parameter" in stripped:
                    found_param = True
                    param_lines.append(stripped)
            else:
                param_lines.append(stripped)
                if ")" in stripped:
                    break

    path.param = " ".join(param_lines).replace("\n"," ").replace("\r"," ")
    path.param = path.param[len("parameter"):].lstrip()
    path.param = path.param.split(")")[0].rstrip()
    pms_info(f"Parameter of top design: {path.param}")

    path.port_define = get_port_define(path.rtl_path+"/"+top_file)

    path.sub_path_list,path.sub_rtl_list,path.sub_rtl_list_raw = get_sub_rtl(sheet)
    pms_info(f"RTL files list: {path.rtl_list} {path.sub_rtl_list}")

    path.tb_path = sheet['B4'].value
    path.tb_path = re.sub(r"\$\{top_path\}",path.top_path,path.tb_path,flags=re.IGNORECASE)

    path.cons_path = sheet['B5'].value
    path.cons_path = re.sub(r"\$\{top_path\}",path.top_path,path.cons_path,flags=re.IGNORECASE)

    path.ip_path = sheet['B6'].value
    path.ip_path = re.sub(r"\$\{top_path\}",path.top_path,path.cons_path,flags=re.IGNORECASE)

    path.mem_path = sheet['B7'].value
    path.mem_path = re.sub(r"\$\{top_path\}",path.top_path,path.mem_path,flags=re.IGNORECASE)

    path.lib_path = sheet['B8'].value
    path.lib_path = re.sub(r"\$\{top_path\}",path.top_path,path.lib_path,flags=re.IGNORECASE)

    mem_list = get_all_mem(path.mem_path,path.sub_path_list,sheet)
    print(mem_list)

    pms_info(f"IP dir: {path.ip_path}")
    pms_info(f"Memory dir: {path.mem_path}")
    pms_info(f"Library dir: {path.lib_path}")
    print()

    return path

def parse_config(sheet):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Fetch configure information for {sheet.title}")
    pms_msg("#"+"-"*60)
    config=Config()
    config.path_root                    = sheet['B2'].value
    config.path_sub                     = standardize_list(sheet['C2'].value)
    if sheet.title in ('sim'):
        config.library.post_sim             = sheet['B5'].value
    if sheet.title in ('synth','sta'):
        config.library.target               = standardize_list(sheet['B5'].value)
        config.library.link                 = standardize_list(sheet['C5'].value)
        config.library.symbol               = sheet['D5'].value
        config.library.synthetic            = sheet['E5'].value
        config.signal_driving_cell.name     = sheet['B8'].value
        config.signal_driving_cell.output   = sheet['D8'].value
        config.signal_driving_cell.input    = sheet['E8'].value
        config.signal_driving_cell.lib      = sheet['F8'].value
        config.clock_driving_cell.name      = sheet['B9'].value
        config.clock_driving_cell.output    = sheet['D9'].value
        config.clock_driving_cell.input     = sheet['E9'].value
        config.clock_driving_cell.lib       = sheet['F9'].value
        config.signal_trans.min             = sheet['B12'].value
        config.signal_trans.max             = sheet['C12'].value
        config.output_load.min              = sheet['B13'].value
        config.output_load.max              = sheet['C13'].value
        config.output_trans.min             = sheet['B14'].value
        config.output_trans.max             = sheet['C14'].value
        config.output_delay.min             = sheet['B15'].value
        config.output_delay.max             = sheet['C15'].value
        config.input_delay.min              = sheet['B16'].value
        config.input_delay.max              = sheet['C16'].value
        config.input_trans.min              = sheet['B17'].value
        config.input_trans.max              = sheet['C17'].value
        config.wire_load_model              = sheet['F12'].value
        config.opera_condition              = sheet['F15'].value
        config.setup_margin                 = sheet['B19'].value
        config.hold_margin                  = sheet['B20'].value
        config.ct.sm.source_latency.min     = sheet['B25'].value
        config.ct.sm.source_latency.max     = sheet['C25'].value
        config.ct.sm.network_latency.min    = sheet['D25'].value
        config.ct.sm.network_latency.max    = sheet['E25'].value
        config.ct.sm.trans.min              = sheet['F25'].value
        config.ct.sm.trans.max              = sheet['G25'].value
        config.ct.sm.skew                   = sheet['H25'].value
        config.ct.sm.noise                  = sheet['I25'].value
        config.ct.md.source_latency.min     = sheet['B26'].value
        config.ct.md.source_latency.max     = sheet['C26'].value
        config.ct.md.network_latency.min    = sheet['D26'].value
        config.ct.md.network_latency.max    = sheet['E26'].value
        config.ct.md.trans.min              = sheet['F26'].value
        config.ct.md.trans.max              = sheet['G26'].value
        config.ct.md.skew                   = sheet['H26'].value
        config.ct.md.noise                  = sheet['I26'].value
        config.ct.lg.source_latency.min     = sheet['B27'].value
        config.ct.lg.source_latency.max     = sheet['C27'].value
        config.ct.lg.network_latency.min    = sheet['D27'].value
        config.ct.lg.network_latency.max    = sheet['E27'].value
        config.ct.lg.trans.min              = sheet['F27'].value
        config.ct.lg.trans.max              = sheet['G27'].value
        config.ct.lg.skew                   = sheet['H27'].value
        config.ct.lg.noise                  = sheet['I27'].value
        config.ct.raw.source_latency.min    = sheet['B28'].value
        config.ct.raw.source_latency.max    = sheet['C28'].value
        config.ct.raw.network_latency.min   = sheet['D28'].value
        config.ct.raw.network_latency.max   = sheet['E28'].value
        config.ct.raw.trans.min             = sheet['F28'].value
        config.ct.raw.trans.max             = sheet['G28'].value
        config.ct.raw.skew                  = sheet['H28'].value
        config.ct.raw.noise                 = sheet['I28'].value

        if config.library.target is None or str(config.library.target).strip() == "":
            pms_fatal("Not define target library")
        config.library.link = config.library.target + " " + config.library.link + " *"
        if config.library.symbol is None or str(config.library.symbol).strip() == "":
            pms_fatal("Not define symbol library")
    for field in fields(config):
        value = getattr(config,field.name)
        pms_info(f"{field.name}: {value}")
    return config

def parse_clock(sheet):
    pms_msg("#"+"-"*60)
    pms_msg("# Fetch clock infomation")
    pms_msg("#"+"-"*60)

    clock_dict  = {}
    row_index   = 4
    columns     = ['level','group','type','period','name','master','jsrc','jmn','jdc','root','add','comment']
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        else:
            row_index += 1
            if re.match(r'^#',row[0]):
                continue
            else:
                row_data = dict(zip(columns,row[:12]))
                name = row_data['name']
                if name:
                    clock_dict[name] = row_data
        
    pms_info(f"Number of clock is {len(clock_dict)}")
    for key,value in clock_dict.items():
        pms_info(f"{key}: {value}")

    return clock_dict
def parse_rst(sheet):
    pms_msg("#"+"-"*60)
    pms_msg("# Fetch reset infomation")
    pms_msg("#"+"-"*60)
    rst_dict  = {}
    row_index   = 3
    columns     = ['level','reset','type','edge','clock']
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        else:
            row_index += 1
            if re.match(r'^#',row[0]):
                continue
            else:
                row_data = dict(zip(columns,row[:5]))
                reset = row_data['reset']
                if reset:
                    rst_dict[reset] = row_data

    pms_info(f"Number of reset is {len(rst_dict)}")
    for key,value in rst_dict.items():
        pms_info(f"{key}: {value}")
    return rst_dict
def parse_io(sheet):
    pms_msg("#"+"-"*60)
    pms_msg("# Fetch IO infomation")
    pms_msg("#"+"-"*60)
    io_dict     = {}
    row_index   = 4
    columns     = ['level','pin','direction','clock','trans_min','trans_max','delay_min','delay_max','delay_cmd','load_min','load_max']
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        else:
            row_index += 1
            if re.match(r'^#',row[0]):
                continue
            else:
                row_data = dict(zip(columns,row[:11]))
                pin = row_data['pin']
                if pin:
                    io_dict[pin] = row_data

    pms_info(f"Number of IO is {len(io_dict)}")
    for key,value in io_dict.items():
        pms_info(f"{key}: {value}")
    return io_dict
def gen_env_sg(path,sg_config):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate environment for SPYGLASS.")
    pms_msg("#"+"-"*60)
    date=datetime.now().strftime("%Y-%m-%d")
    sg_config.path_root = os.path.join(path.top_path,sg_config.path_root)
    script_dir=os.path.dirname(os.path.abspath(__file__))
    makefile_temp = script_dir + "/templates/spyglass/Makefile"
    sg_cons_temp = script_dir + "/templates/spyglass/sg.tcl"
    sub_dir = sg_config.path_sub.split()
    makefile = os.path.join(sg_config.path_root,"Makefile")
    sg_cons = os.path.join(sg_config.path_root,'scripts',f"{path.top}.tcl")
    
    if check_dir(sg_config.path_root,sub_dir) == 0:
        return 0
    
    shutil.copy(makefile_temp,makefile)
    shutil.copy(sg_cons_temp,sg_cons)
    replace_in_file(makefile,'__TOP__',path.top)
    replace_in_file(makefile,'__DATE__',date)
    replace_in_file(makefile,'__ROOT_PATH__',os.path.join(sg_config.path_root))

    replace_in_file(sg_cons,'__TOP__',path.top)
    replace_in_file(sg_cons,'__DATE__',date)
    replace_in_file(sg_cons,'__ROOT_PATH__',sg_config.path_root)


    filelist = "//Src file\n"
    filelist += '\n'.join([path.rtl_path+ '/'+ rtl_file for rtl_file in path.rtl_list.split()])
    filelist += '\n'+"\n".join(path.sub_rtl_list_raw.split())

    with open(os.path.join(path.top_path,sg_config.path_root,"work","filelist.f"),"w",encoding="utf-8")as f:
        f.write(filelist)

    pms_info(f"Please find Makefile in {makefile}")
    pms_info(f"Please find script in {sg_cons}")
    pms_info(f"Please find filelist in {os.path.join(path.top_path,sg_config.path_root,'work','filelist.f')}")
    return 1


def gen_env_sta(path,synth_config,sta_config):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate environment for STA.")
    pms_msg("#"+"-"*60)
    date=datetime.now().strftime("%Y-%m-%d")
    param_netlist = re.sub(r'\s*,\s*','_',path.param)
    param_netlist = re.sub(r'[^\w]','',param_netlist)
    sta_config.path_root = os.path.join(path.top_path,sta_config.path_root)
    script_dir=os.path.dirname(os.path.abspath(__file__))
    sta_cons_temp = script_dir + "/templates/sta/sta.tcl"
    synopsys_setup_temp = script_dir + "/templates/sta/.synopsys_pt.setup"
    makefile_temp = script_dir + "/templates/sta/Makefile"
    sub_dir = sta_config.path_sub.split()

    top_cons = os.path.join(sta_config.path_root,"scripts",path.top+".tcl")
    synopsys_setup = os.path.join(sta_config.path_root,"work",".synopsys_pt.setup")
    makefile = os.path.join(sta_config.path_root,"Makefile")
    if check_dir(sta_config.path_root,sub_dir) == 0:
        return 0

    shutil.copy(sta_cons_temp,top_cons)
    shutil.copy(synopsys_setup_temp,synopsys_setup)
    shutil.copy(makefile_temp,makefile)


    budget = "array set ct_budget {"
    budget += f"\n    sm.source_latency.min    {sta_config.ct.sm.source_latency.min}"
    budget += f"\n    sm.source_latency.max    {sta_config.ct.sm.source_latency.max}"
    budget += f"\n    sm.network_latency.min   {sta_config.ct.sm.network_latency.min}"
    budget += f"\n    sm.network_latency.max   {sta_config.ct.sm.network_latency.max}"
    budget += f"\n    sm.trans.min             {sta_config.ct.sm.trans.min}"
    budget += f"\n    sm.trans.max             {sta_config.ct.sm.trans.max}"
    budget += f"\n    sm.skew                  {sta_config.ct.sm.skew}"
    budget += f"\n    sm.noise                 {sta_config.ct.sm.noise}"
    budget += f"\n    md.source_latency.min    {sta_config.ct.md.source_latency.min}"
    budget += f"\n    md.source_latency.max    {sta_config.ct.md.source_latency.max}"
    budget += f"\n    md.network_latency.min   {sta_config.ct.md.network_latency.min}"
    budget += f"\n    md.network_latency.max   {sta_config.ct.md.network_latency.max}"
    budget += f"\n    md.trans.min             {sta_config.ct.md.trans.min}"
    budget += f"\n    md.trans.max             {sta_config.ct.md.trans.max}"
    budget += f"\n    md.skew                  {sta_config.ct.md.skew}"
    budget += f"\n    md.noise                 {sta_config.ct.md.noise}"
    budget += f"\n    lg.source_latency.min    {sta_config.ct.lg.source_latency.min}"
    budget += f"\n    lg.source_latency.max    {sta_config.ct.lg.source_latency.max}"
    budget += f"\n    lg.network_latency.min   {sta_config.ct.lg.network_latency.min}"
    budget += f"\n    lg.network_latency.max   {sta_config.ct.lg.network_latency.max}"
    budget += f"\n    lg.trans.min             {sta_config.ct.lg.trans.min}"
    budget += f"\n    lg.trans.max             {sta_config.ct.lg.trans.max}"
    budget += f"\n    lg.skew                  {sta_config.ct.lg.skew}"
    budget += f"\n    lg.noise                 {sta_config.ct.lg.noise}"
    budget += f"\n    raw.source_latency.min   {sta_config.ct.raw.source_latency.min}"
    budget += f"\n    raw.source_latency.max   {sta_config.ct.raw.source_latency.max}"
    budget += f"\n    raw.network_latency.min  {sta_config.ct.raw.network_latency.min}"
    budget += f"\n    raw.network_latency.max  {sta_config.ct.raw.network_latency.max}"
    budget += f"\n    raw.trans.min            {sta_config.ct.raw.trans.min}"
    budget += f"\n    raw.trans.max            {sta_config.ct.raw.trans.max}"
    budget += f"\n    raw.skew                 {sta_config.ct.raw.skew}"
    budget += f"\n    raw.noise                {sta_config.ct.raw.noise}"
    budget += "\n}"
    budget += f"\nset SETUP_MARGIN  {sta_config.setup_margin}"
    budget += f"\nset HOLD_MARGIN   {sta_config.hold_margin}"


    replace_in_file(makefile,'__SCRIPT__',top_cons)

    replace_in_file(synopsys_setup,'__ROOT__',path.top_path)
    replace_in_file(synopsys_setup,'__LIB_PATH__',path.lib_path)
    replace_in_file(synopsys_setup,'__NETLIST_PATH__',synth_config.path_root+'/mapped')
    replace_in_file(synopsys_setup,'__TARGET_LIB__',synth_config.library.target)
    replace_in_file(synopsys_setup,'__LINK_LIB__',synth_config.library.link)
    replace_in_file(top_cons,'__DATE__',date)
    replace_in_file(top_cons,'__TOP__',path.top)
    replace_in_file(top_cons,'__TARGET_LIB__',synth_config.library.target)
    replace_in_file(top_cons,'__TOP_NETLIST_MODULE__',path.top+'_'+param_netlist)
    replace_in_file(top_cons,'__NETLIST__',path.top+'.v')
    replace_in_file(top_cons,'__CLOCK_LIB_NAME__',synth_config.clock_driving_cell.lib)
    replace_in_file(top_cons,'__CLOCK_DRIVE_CELL__',synth_config.clock_driving_cell.name)
    replace_in_file(top_cons,'__CLOCK_DRIVE_PIN__',synth_config.clock_driving_cell.output)
    replace_in_file(top_cons,'__SIGNAL_LIB_NAME__',synth_config.signal_driving_cell.lib)
    replace_in_file(top_cons,'__SIGNAL_DRIVE_CELL__',synth_config.signal_driving_cell.name)
    replace_in_file(top_cons,'__SIGNAL_DRIVE_PIN__',synth_config.signal_driving_cell.output)
    replace_in_file(top_cons,'__WIRE_LOAD_MODEL__',synth_config.wire_load_model)
    replace_in_file(top_cons,'__OPERA_CONDITION__',synth_config.opera_condition)
    replace_in_file(top_cons,'__CLOCK_TREE__BUDGET__',budget)

    replace_in_file(top_cons,'__CLOCKS_CONSTRAINT__',os.path.join(path.cons_path,path.top+"_clk.tcl"))
    replace_in_file(top_cons,'__RESET_CONSTRAINT__',os.path.join(path.cons_path,path.top+"_rst.tcl"))
    replace_in_file(top_cons,'__IO_CONSTRAINT__',os.path.join(path.cons_path,path.top+"_io.tcl"))

def gen_env_synth(path,synth_config):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate environment for SYNTH.")
    pms_msg("#"+"-"*60)
    date=datetime.now().strftime("%Y-%m-%d")
    synth_config.path_root = os.path.join(path.top_path,synth_config.path_root)
    script_dir=os.path.dirname(os.path.abspath(__file__))
    synth_cons_temp = script_dir + "/templates/synthesis/synth.tcl"
    synopsys_setup_temp = script_dir + "/templates/synthesis/.synopsys_dc.setup"
    makefile_temp = script_dir + "/templates/synthesis/Makefile"
    sub_dir = synth_config.path_sub.split()

    top_cons = os.path.join(synth_config.path_root,"scripts",path.top+".tcl")
    synopsys_setup = os.path.join(synth_config.path_root,"work",".synopsys_dc.setup")
    makefile = os.path.join(synth_config.path_root,"Makefile")
    if check_dir(synth_config.path_root,sub_dir) == 0:
        return 0

    shutil.copy(synth_cons_temp,top_cons)
    shutil.copy(synopsys_setup_temp,synopsys_setup)
    shutil.copy(makefile_temp,makefile)

    budget = "array set ct_budget {"
    budget += f"\n    sm.source_latency.min    {synth_config.ct.sm.source_latency.min}"
    budget += f"\n    sm.source_latency.max    {synth_config.ct.sm.source_latency.max}"
    budget += f"\n    sm.network_latency.min   {synth_config.ct.sm.network_latency.min}"
    budget += f"\n    sm.network_latency.max   {synth_config.ct.sm.network_latency.max}"
    budget += f"\n    sm.trans.min             {synth_config.ct.sm.trans.min}"
    budget += f"\n    sm.trans.max             {synth_config.ct.sm.trans.max}"
    budget += f"\n    sm.skew                  {synth_config.ct.sm.skew}"
    budget += f"\n    sm.noise                 {synth_config.ct.sm.noise}"
    budget += f"\n    md.source_latency.min    {synth_config.ct.md.source_latency.min}"
    budget += f"\n    md.source_latency.max    {synth_config.ct.md.source_latency.max}"
    budget += f"\n    md.network_latency.min   {synth_config.ct.md.network_latency.min}"
    budget += f"\n    md.network_latency.max   {synth_config.ct.md.network_latency.max}"
    budget += f"\n    md.trans.min             {synth_config.ct.md.trans.min}"
    budget += f"\n    md.trans.max             {synth_config.ct.md.trans.max}"
    budget += f"\n    md.skew                  {synth_config.ct.md.skew}"
    budget += f"\n    md.noise                 {synth_config.ct.md.noise}"
    budget += f"\n    lg.source_latency.min    {synth_config.ct.lg.source_latency.min}"
    budget += f"\n    lg.source_latency.max    {synth_config.ct.lg.source_latency.max}"
    budget += f"\n    lg.network_latency.min   {synth_config.ct.lg.network_latency.min}"
    budget += f"\n    lg.network_latency.max   {synth_config.ct.lg.network_latency.max}"
    budget += f"\n    lg.trans.min             {synth_config.ct.lg.trans.min}"
    budget += f"\n    lg.trans.max             {synth_config.ct.lg.trans.max}"
    budget += f"\n    lg.skew                  {synth_config.ct.lg.skew}"
    budget += f"\n    lg.noise                 {synth_config.ct.lg.noise}"
    budget += f"\n    raw.source_latency.min   {synth_config.ct.raw.source_latency.min}"
    budget += f"\n    raw.source_latency.max   {synth_config.ct.raw.source_latency.max}"
    budget += f"\n    raw.network_latency.min  {synth_config.ct.raw.network_latency.min}"
    budget += f"\n    raw.network_latency.max  {synth_config.ct.raw.network_latency.max}"
    budget += f"\n    raw.trans.min            {synth_config.ct.raw.trans.min}"
    budget += f"\n    raw.trans.max            {synth_config.ct.raw.trans.max}"
    budget += f"\n    raw.skew                 {synth_config.ct.raw.skew}"
    budget += f"\n    raw.noise                {synth_config.ct.raw.noise}"
    budget += "\n}"
    budget += f"\nset SETUP_MARGIN  {synth_config.setup_margin}"
    budget += f"\nset HOLD_MARGIN   {synth_config.hold_margin}"


    replace_in_file(makefile,'__SCRIPT__',top_cons)

    replace_in_file(synopsys_setup,'__ROOT__',path.top_path)
    replace_in_file(synopsys_setup,'__LIB_PATH__',path.lib_path)
    replace_in_file(synopsys_setup,'__RTL_PATH__',path.rtl_path+path.sub_path_list)
    replace_in_file(synopsys_setup,'__TARGET_LIB__',synth_config.library.target)
    replace_in_file(synopsys_setup,'__LINK_LIB__',synth_config.library.link)
    replace_in_file(synopsys_setup,'__SYMBOL_LIB__',synth_config.library.symbol)
    replace_in_file(synopsys_setup,'__SYNTH_LIB__',synth_config.library.synthetic)
    replace_in_file(top_cons,'__DATE__',date)
    replace_in_file(top_cons,'__TOP__',path.top)
    replace_in_file(top_cons,'__TARGET_LIB__',synth_config.library.target)
    replace_in_file(top_cons,'__RTL_LIST__',path.rtl_list+path.sub_rtl_list)
    replace_in_file(top_cons,'__PARAMETER__',path.param)
    replace_in_file(top_cons,'__TARGET_LIB__',synth_config.library.target)
    replace_in_file(top_cons,'__CLOCK_LIB_NAME__',synth_config.clock_driving_cell.lib)
    replace_in_file(top_cons,'__CLOCK_DRIVE_CELL__',synth_config.clock_driving_cell.name)
    replace_in_file(top_cons,'__CLOCK_DRIVE_PIN__',synth_config.clock_driving_cell.output)
    replace_in_file(top_cons,'__SIGNAL_LIB_NAME__',synth_config.signal_driving_cell.lib)
    replace_in_file(top_cons,'__SIGNAL_DRIVE_CELL__',synth_config.signal_driving_cell.name)
    replace_in_file(top_cons,'__SIGNAL_DRIVE_PIN__',synth_config.signal_driving_cell.output)
    replace_in_file(top_cons,'__WIRE_LOAD_MODEL__',synth_config.wire_load_model)
    replace_in_file(top_cons,'__OPERA_CONDITION__',synth_config.opera_condition)
    replace_in_file(top_cons,'__CLOCK_TREE__BUDGET__',budget)

    replace_in_file(top_cons,'__CLOCKS_CONSTRAINT__',os.path.join(path.cons_path,path.top+"_clk.tcl"))
    replace_in_file(top_cons,'__RESET_CONSTRAINT__',os.path.join(path.cons_path,path.top+"_rst.tcl"))
    replace_in_file(top_cons,'__IO_CONSTRAINT__',os.path.join(path.cons_path,path.top+"_io.tcl"))

def gen_env_sim(path,synth_config,sim_config):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate environment for SIMULATION.")
    pms_msg("#"+"-"*60)
    date=datetime.now().strftime("%Y-%m-%d")
    sim_config.path_root = os.path.join(path.top_path,sim_config.path_root)
    script_dir=os.path.dirname(os.path.abspath(__file__))
    makefile_temp = script_dir + "/templates/simulation/Makefile"
    sub_dir = sim_config.path_sub.split()
    makefile = os.path.join(sim_config.path_root,"Makefile")
    if check_dir(sim_config.path_root,sub_dir) == 0:
        return 0

    shutil.copy(makefile_temp,makefile)

    replace_in_file(makefile,'__TOP__',path.top)
    replace_in_file(makefile,'__DATE__',date)
    replace_in_file(makefile,'__ROOT_PATH__',sim_config.path_root)
    replace_in_file(makefile,'__POST_SIM_LIB__',os.path.join(path.lib_path,'verilog',sim_config.library.post_sim))

    filelist = "//Src file\n"
    filelist += '\n'.join([path.rtl_path+ '/'+ rtl_file for rtl_file in path.rtl_list.split()])
    filelist += '\n'+"\n".join(path.sub_rtl_list_raw.split())
    filelist += f"\n//Testbench file\n{path.tb_path}/{path.top}_tb.sv"
    with open(os.path.join(path.top_path,sim_config.path_root,"work","filelist.f"),"w",encoding="utf-8")as f:
        f.write(filelist)
    filelist_post = f"//Gate netlist\n{synth_config.path_root}/mapped/{path.top}.v"
    filelist_post += f"\n//Testbench file\n{path.tb_path}/{path.top}_tb.sv"
    with open(os.path.join(path.top_path,sim_config.path_root,"work","filelist_post.f"),"w",encoding="utf-8")as f:
        f.write(filelist_post)

def gen_cons_clk(path,clock_dict):
    if not clock_dict:
        pms_msg("No clock data found.")
        return
    cons = "\n#========================================\n#Clock Constraint"

    for name,row_data in clock_dict.items():
        cons += f"\n#Clock {name} - {row_data['comment']}"
        if clock_dict[name]['root'] is None:
            # is virtual clock
            cons += f"\n#----------------------------------------\ncreate_clock -name {name} -period {clock_dict[name]['period']}"
        else:
            if "/" in clock_dict[name]['root']:
                pin_or_port = "pins"
            else:
                pin_or_port = "ports"
            
            if clock_dict[name]['master'] is None:
                # is master clock
                cons += f"\nset JITTER [expr {clock_dict[name]['jsrc']} + {clock_dict[name]['jmn']} + {clock_dict[name]['jdc']}]"
                cons += f"\nset SETUP_UNCERTAINTY [expr $ct_budget({clock_dict[name]['type']}.skew) + $ct_budget({clock_dict[name]['type']}.noise) + $JITTER + ({clock_dict[name]['period']} - $ct_budget({clock_dict[name]['type']}.skew) - $ct_budget({clock_dict[name]['type']}.noise) - $JITTER) * $SETUP_MARGIN]"
                cons += f"\nset HOLD_UNCERTAINTY [expr $ct_budget({clock_dict[name]['type']}.skew) + $ct_budget({clock_dict[name]['type']}.noise) + $HOLD_MARGIN]"
                cons += f"\ncreate_clock -name {name} -period {clock_dict[name]['period']} [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_ideal_network [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_dont_touch_network [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_drive 0 [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_clock_uncertainty  -setup $SETUP_UNCERTAINTY [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nset_clock_uncertainty  -hold $HOLD_UNCERTAINTY [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nset_clock_transition  -max $ct_budget({clock_dict[name]['type']}.trans.max) [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nset_clock_transition  -min $ct_budget({clock_dict[name]['type']}.trans.min) [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nset_clock_latency -source -max $ct_budget({clock_dict[name]['type']}.source_latency.max) [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nset_clock_latency -source -min $ct_budget({clock_dict[name]['type']}.source_latency.min) [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nset_clock_latency -max $ct_budget({clock_dict[name]['type']}.network_latency.max) [get_clocks {clock_dict[name]['root']}]"
                cons += f"\nif {{$ct_budget({clock_dict[name]['type']}.network_latency.min) ne \"NA\"}} {{"
                cons += f"\n    set_clock_latency -min $ct_budget({clock_dict[name]['type']}.network_latency.min) [get_clocks {clock_dict[name]['root']}]"
                cons += f"\n}}"
            else:
                # is generated clock
                if clock_dict[name]['master'] in clock_dict:
                    mst_source = clock_dict[clock_dict[name]['master']]['root']
                else:
                    mst_source = clock_dict[name]['master']
                print("1111")

                if not "/" in mst_source:
                    mst_source = f"[get_port {mst_source}]"
                else:
                    mst_source = f"[get_pins {mst_source}]"

                if match := re.match(r"-div\s+(\d+)$",clock_dict[name]['period']):
                    generated_period = f"-divide_by {match.group(1)}"
                elif match := re.match(r"-multi\s+(\d+)$",clock_dict[name]['period']):
                    generated_period = f"-multiply_by {match.group(1)}"
                elif match := re.match(r"-edges\s+\{\s*\d+\s+\d+\s+\d+\s*\}$",clock_dict[name]['period']):
                    generated_period = clock_dict[name]['period']
                else:
                    generated_period = clock_dict[name]['period']
                
                cons += f"\ncreate_generated_clock -name {name} {generated_period} -source {mst_source} [get_{pin_or_port} {clock_dict[name]['root']}]" 
                if  clock_dict[name]['add'] is not None:
                    if not "/" in clock_dict[name]['add']:
                        cons += f"\ncreate_generated_clock -name {name} {generated_period} -source {mst_source} -add [get_port {clock_dict[name]['add']}]" 
                    else:
                        pms_error(f"Add object is not a port: {clock_dict[name]['add']}")

    cons += f"\n#----------------------------------------\n#Set var"

    grouped_clks = {}
    for clk_info in clock_dict.values():
        group = clk_info['group']
        name = clk_info['name']
        grouped_clks.setdefault(group,[]).append(name)
    
    cons += "\n#----------------------------------------\n#Set clock groups"
    cons += "\nset_clock_groups -asynchronous "
    for group,names in grouped_clks.items():
        cons += " -group {"
        cons += f"{' '.join(names)}"
        cons += "}"

    cons_file = os.path.join(path.cons_path,path.top+"_clk.tcl")
    with open(cons_file,"w",encoding="utf-8")as f:
        f.write(cons)
    pms_info(f"Please find the clock constraint in {cons_file}")


def gen_cons_rst(path,rst_dict,clock_dict):
    if not rst_dict:
        pms_msg("No Reset data found.")
        return

    cons = "\n#========================================\n#Reset Constraint"
    
    for reset,row_data in rst_dict.items():
        cons += f"\nset_ideal_network [get_port {reset}]"
        cons += f"\nset_dont_touch_network [get_port {reset}]"
        cons += f"\nset_drive 0 [get_port {reset}]"
        cons += f"\nset_false_path -from [get_port {reset}]"
    
    cons_file = os.path.join(path.cons_path,path.top+"_rst.tcl")
    with open(cons_file,"w",encoding="utf-8")as f:
        f.write(cons)
    pms_info(f"Please find the reset constranit in {cons_file}")

def gen_cons_io(path,io_dict,clock_dict):
    if not io_dict:
        pms_msg("No IO data found.")
        return
    input_ports_except_clk = ""
    output_ports_except_clk = ""
    cons = "\n#========================================\n#IO Constraint"
    
    #for row in enumerate(clock_list,start=1):
    for pin,row_data in io_dict.items():
        if io_dict[pin]['direction'] in ("input","in"):
            input_ports_except_clk += f" {pin}"
            if io_dict[pin]['delay_min'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_min']
                cons += f"\nset_input_delay -clock {io_dict[pin]['clock']} -min {delay} [get_ports {pin}]"
            if io_dict[pin]['delay_max'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_max']
                cons += f"\nset_input_delay -clock {io_dict[pin]['clock']} -max {delay} [get_ports {pin}]"
        elif io_dict[pin]['direction'] in ("output","out"):
            output_ports_except_clk += f" {pin}"
            if io_dict[pin]['delay_min'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_min']
                cons += f"\nset_output_delay -clock {io_dict[pin]['clock']} -min {delay} [get_ports {pin}]"
            if io_dict[pin]['delay_max'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_max']
                cons += f"\nset_output_delay -clock {io_dict[pin]['clock']} -max {delay} [get_ports {pin}]"
            if io_dict[pin]['load_min'] is not None:
                load = io_dict[pin]['load_min']
                cons += f"\nset_load -min {load} [get_ports {pin}]"
            if io_dict[pin]['load_max'] is not None:
                load = io_dict[pin]['load_max']
                cons += f"\nset_load -max {load} [get_ports {pin}]"

    cons += "\n#----------------------------------------\n#Set var"
    cons += f"\nset NON_CLK_INPUT_PORTS [get_ports -quiet \"{input_ports_except_clk}\"]"
    cons += f"\nset NON_CLK_OUTPUT_PORTS [get_ports -quiet \"{output_ports_except_clk}\"]"
    cons += "\n#----------------------------------------\n#Set input"
    cons += "\nset_driving_cell -lib_cell ${SIGNAL_DRIVE_CELL} -pin ${SIGNAL_DRIVE_PIN} -library ${SIGNAL_LIB_NAME} ${NON_CLK_INPUT_PORTS}"
    cons += "\n#----------------------------------------\n#Set output"
    cons += "\n"
    cons += "\n#----------------------------------------\n#Set false path"
    cons += "\nset_false_path -from ${NON_CLK_INPUT_PORTS} -thr ${NON_CLK_OUTPUT_PORTS}"

    cons_file = os.path.join(path.cons_path,path.top+"_io.tcl")
    with open(cons_file,"w",encoding="utf-8")as f:
        f.write(cons)
    pms_info(f"Please find the IO constranit in {cons_file}")

def gen_sdc(path,clock_dict,rst_dict,io_dict):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate SDC")
    pms_msg("#"+"-"*60)
    pms_info(f"Generate clock constraints for {len(clock_dict)} clocks.")
    gen_cons_clk(path,clock_dict)
    pms_info(f"Generate reset constraints for {len(rst_dict)} resets.")
    gen_cons_rst(path,rst_dict,clock_dict)
    pms_info(f"Generate IO constraints for {len(io_dict)} ports.")
    gen_cons_io(path,io_dict,clock_dict)

def gen_cn(clock_dict,rst_dict,io_dict,path,sg_config):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate constranit for SPYGLASS.")
    pms_msg("#"+"-"*60)
    date=datetime.now().strftime("%Y-%m-%d")
    awl = f"#{date}\n#========================================\n#Waiver"
    awl += f"\nwaive -du {{  {{{path.top}}}  }}  -msg {{\'timeunit\' construct is not synthesizable. Ignoring for synthesis}}  -rule {{  {{SYNTH_78}}  }}" 
    awl += f"\nwaive -du {{  {{{path.top}}}  }}  -msg {{\'timeprecision\' construct is not synthesizable. Ignoring for synthesis}}  -rule {{  {{SYNTH_78}}  }}"
    awl += f"\nwaive -du {{  {{{path.top}}}  }}  -msg {{Delay used without timescale compiler directive}}  -rule {{  {{CheckDelayTimescale-ML}}  }} "
    sgdc = f"#{date}\n#========================================\n#SGDC"
    sgdc += f"\ncurrent_design	{path.top}"
    sgdc += f"\n#Clock"
    for name,row_data in clock_dict.items():
        #if clock_dict[name]['root'] and "/" not in clock_dict[name]['root']:
        if clock_dict[name]['root']:
            if isinstance(clock_dict[name]['period'],(float,int)):
                sgdc += f"\nclock -name {name} -domain {clock_dict[name]['group']} -edge {{\"0\" \"{clock_dict[name]['period']/2}\"}} -period {clock_dict[name]['period']}"
            elif match := re.match(r"-div\S*\s+(\d+)$",clock_dict[name]['period']):
                generated_period = float(clock_dict[clock_dict[name]['master']]['period']) * float(match.group(1))
                sgdc += f"\nclock -name {name} -domain {clock_dict[name]['group']} -edge {{\"0\" \"{generated_period/2}\"}} -period {generated_period}"
            elif match := re.match(r"-multi\s+(\d+)$",clock_dict[name]['period']):
                generated_period = float(clock_dict[clock_dict[name]['master']]['period']) / (match.group(1))
                sgdc += f"\nclock -name {name} -domain {clock_dict[name]['group']} -edge {{\"0\" \"{generated_period/2}\"}} -period {generated_period}"
            elif match := re.match(r"-edges\s+\{\s*\d+\s+\d+\s+\d+\s*\}$",clock_dict[name]['period']):
                pms_fatal("Please fix script to support constraiting clock with waveform for SGDC")
            else:
                pms_fatal(f"Can not capture the type of period for clock: {name}")

    sgdc += f"\n#Reset"
    for reset,row_data in rst_dict.items():
        if "neg" in rst_dict[reset]['edge']:
            sgdc += f"\nreset -name {reset} -{rst_dict[reset]['type']} -value 0"
        elif "pos" in rst_dict[reset]['edge']:
            sgdc += f"\nreset -name {reset} -{rst_dict[reset]['type']} -value 1"
        else:
            pms_fatal(f"Can not capture value of reset {reset}")
    with open(os.path.join(sg_config.path_root,'cn',path.top+'.awl'),"w",encoding="utf-8")as f:
        f.write(awl)
    pms_info(f"Please find awl in {os.path.join(sg_config.path_root,'cn',path.top+'.awl')}")
    with open(os.path.join(sg_config.path_root,'cn',path.top+'.sgdc'),"w",encoding="utf-8")as f:
        f.write(sgdc)
    pms_info(f"Please find sgdc in {os.path.join(sg_config.path_root,'cn',path.top+'.sgdc')}")

def gen_tb(clock_dict,rst_dict,io_dict,path,sim_config,synth_config):
    pms_msg("#"+"-"*60)
    pms_msg(f"# Generate testbench for SIMULATION.")
    pms_msg("#"+"-"*60)
    date=datetime.now().strftime("%Y-%m-%d")
    if check_dir(path.tb_path) == 0:
        return 0

    matches = re.findall(r'(\w+)\s*=',path.param)
    param_tb = ','.join([f'.{name}({name})' for name in matches])
    param_netlist = re.sub(r'\s*,\s*','_',path.param)
    param_netlist = re.sub(r'[^\w]','',param_netlist)


    tb = f"//{date}\n//========================================\n//The time unit and precision"
    tb += "\n`timescale  1ns/1ps"
    tb += f"\nmodule {path.top}_tb;"
    tb += "\n    //========================================"
    tb += f"\n    parameter {path.param};"
    for name,row_data in clock_dict.items():
        if clock_dict[name]['root'] and "/" not in clock_dict[name]['root']:
            tb += f"\n    parameter PERIOD_{name.upper()} = {clock_dict[name]['period']};"
    tb += "\n"+path.port_define
    tb += "\n    //========================================\n    //Instantiate"
    tb += f"\n    `ifdef POST_SIM"
    tb += f"\n        {path.top}_{param_netlist} u_{path.top}(.*);"
    tb += f"\n    `else"
    tb += f"\n        {path.top} #(\n            {param_tb}\n        ) u_{path.top}(.*);"
    tb += f"\n    `endif"
    tb += "\n    //========================================\n    //Clock drive"
    tb += "\n    initial begin"
    for name,row_data in clock_dict.items():
        if clock_dict[name]['root'] and "/" not in clock_dict[name]['root']:
            tb += f"\n        {name} = '0;\n        forever #(PERIOD_{name.upper()}/2) {name}= ~{name};"
    tb += "\n    end"
    tb += "\n    //========================================\n    //Task reset"
    rst_list = ""
    for reset,row_data in rst_dict.items():
        tb += f"\n    task task_rst_{reset};"
        rst_list += f"        task_rst_{reset};\n"
        if "neg" in rst_dict[reset]['edge']:
            tb += f"\n        {reset} = '0;#1;"
            tb += f"\n        {reset} = '1;#1;"
        elif "pos" in rst_dict[reset]['edge']:
            tb += f"\n        {reset} = '1;#1;"
            tb += f"\n        {reset} = '0;#1;"
        else:
            pms_fatal(f"Can not capture value of reset {reset}")
        tb += f"\n    endtask"
    tb += "\n    //========================================\n    //Task initial"
    tb += "\n    task task_init;"
    for pin,row_data in io_dict.items():
        if "in" in io_dict[pin]['direction']:
            tb += f"\n        {pin} = '0;"
    tb += f"\n        #5;"
    tb += "\n    endtask"
    tb += "\n    //========================================\n    //Simulation"
    tb += "\n    initial begin"
    tb += "\n        //Reset&Init"
    tb += "\n        task_init;"
    tb += f"\n{rst_list}"
    tb += "\n        //Simulation behavior\n\n\n"
    tb += "\n        #400;"
    tb += "\n        $display(\"\\033[31;5m 仿真完成! \\033[0m\",`__FILE__,`__LINE__);"
    tb += "\n        $finish;"
    tb += "\n    end"
    tb += "\n    //========================================\n    //VCS Simulation"
    tb += "\n    `ifdef VCS_SIM"
    tb += "\n        //VCS系统函数"
    tb += "\n        initial begin"
    tb += "\n            $vcdpluson(); //打开VCD+文件记录"
    tb += f"\n            $fsdbDumpfile(\"{os.path.join(path.top_path,sim_config.path_root)}/sim/{path.top}.fsdb\"); //生成fsdb"
    tb += "\n            $fsdbDumpvars(\"+all\");"
    tb += "\n            $vcdplusmemon(); //查看多维数组"
    tb += "\n        end"
    tb += "\n        //后仿真"
    tb += "\n        `ifdef POST_SIM"
    tb += "\n        //back annotate the SDF file"
    tb += "\n        initial begin"
    tb += f"\n            $sdf_annotate(\"{os.path.join(path.top_path,synth_config.path_root,'mapped',path.top+'.sdf')}\","
    tb += f"\n                          {path.top}_tb.u_{path.top},,,"
    tb += f"\n                          \"TYPICAL\","
    tb += f"\n                          \"1:1:1\","
    tb += f"\n                          \"FROM_MTM\");"
    tb += "\n            $display(\"\\033[31;5m back annotate \033[0m\",`__FILE__,`__LINE__);"
    tb += "\n        end"
    tb += "\n        `endif"
    tb += "\n    `endif"
    tb += "\nendmodule"
    with open(os.path.join(path.tb_path,path.top+"_tb.sv"),"w",encoding="utf-8")as f:
        f.write(tb)

def main(filename):
    wb = openpyxl.load_workbook(filename)
    print_startup()

    if 'path' in wb.sheetnames:
        path = parse_path(wb['path'])
    else:
        pms_fatal(f"Can not find \"path\" sheet in {filename}")
    if 'clock' in wb.sheetnames:
        clock_dict = parse_clock(wb['clock'])
    else:
        pms_fatal(f"Can not find \"clock\" sheet in {filename}")
    if 'reset' in wb.sheetnames:
        rst_dict = parse_rst(wb['reset'])
    else:
        pms_fatal(f"Can not find \"reset\" sheet in {filename}")
    if 'io' in wb.sheetnames:
        io_dict = parse_io(wb['io'])
    else:
        pms_fatal(f"Can not find \"io\" sheet in {filename}")
    
    if check_dir(path.cons_path) == 1:
        gen_sdc(path,clock_dict,rst_dict,io_dict)

    if 'spyglass'in wb.sheetnames:
        sg_config = parse_config(wb['spyglass'])
        if gen_env_sg(path,sg_config) == 1:
            gen_cn(clock_dict,rst_dict,io_dict,path,sg_config)
    if 'synth' in wb.sheetnames:
        synth_config = parse_config(wb['synth'])
        gen_env_synth(path,synth_config)
    if 'sta' in wb.sheetnames:
        sta_config = parse_config(wb['sta'])
        gen_env_sta(path,synth_config,sta_config)
    if 'sim' in wb.sheetnames:
        sim_config = parse_config(wb['sim'])
        gen_tb(clock_dict,rst_dict,io_dict,path,sim_config,synth_config)
        gen_env_sim(path,synth_config,sim_config)



    wb.close()

if __name__ == "__main__":
    #if len(sys.argv) != 2:
    #    pms_fatal("Usage: python cms.py <excel_file>")
    
    if len(sys.argv) > 1 and sys.argv[1].strip():
        table_path = sys.argv[1].strip()
    else:
        xls_files = glob.glob(os.path.join(os.getcwd(),'pms','*.xls*'))
        if xls_files:
            table_path = xls_files[0]
        else:
            pms_fatal("Can not find PMS table.")

    try:
        result = main(table_path)
    except Exception as e:
        print("Error:", e)
