import openpyxl
import sys
from dataclasses import dataclass

@dataclass
class DrivingCell:
    name: str = None
    output: str = None
    input: str = None
    lib_name: str = None

@dataclass
class SignalTrans:
    min: float = None
    max: float = None

@dataclass
class CtbSm:
    source: SignalTrans = SignalTrans()
    network: SignalTrans = SignalTrans()
    trans: SignalTrans = SignalTrans()
    skew: float = None
    noise: float = None

@dataclass
class PtSyncData:
    signal_driving_cell: DrivingCell = DrivingCell()
    clock_driving_cell: DrivingCell = DrivingCell()
    signal_trans: SignalTrans = SignalTrans()
    output_load: SignalTrans = SignalTrans()
    setup_margin: float = None
    hold_margin: float = None
    ctb: CtbSm = CtbSm()

def find_cell(sheet, target):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == target:
                return cell
    return None

def parse_pt_sync_sheet(sheet):
    data = PtSyncData()
    
    # Signal Driving Cell
    cell = find_cell(sheet, "Signal Driving Cell")
    if cell:
        data.signal_driving_cell.name = sheet.cell(cell.row, cell.column+1).value
        data.signal_driving_cell.output = sheet.cell(cell.row, cell.column+2).value
        data.signal_driving_cell.input = sheet.cell(cell.row, cell.column+3).value
        data.signal_driving_cell.lib_name = sheet.cell(cell.row, cell.column+4).value
    
    # Clock Driving Cell
    cell = find_cell(sheet, "Clock driving Cell")
    if cell:
        data.clock_driving_cell.name = sheet.cell(cell.row, cell.column+1).value
        data.clock_driving_cell.output = sheet.cell(cell.row, cell.column+2).value
        data.clock_driving_cell.input = sheet.cell(cell.row, cell.column+3).value
        data.clock_driving_cell.lib_name = sheet.cell(cell.row, cell.column+4).value
    
    # Signal Transition
    cell = find_cell(sheet, "Signal Transition")
    if cell:
        data.signal_trans.min = sheet.cell(cell.row, cell.column+1).value
        data.signal_trans.max = sheet.cell(cell.row, cell.column+2).value
    
    # Output Load
    cell = find_cell(sheet, "Output Load")
    if cell:
        data.output_load.min = sheet.cell(cell.row, cell.column+1).value
        data.output_load.max = sheet.cell(cell.row, cell.column+2).value
    
    # Setup Margin
    cell = find_cell(sheet, "Setup Matgin")
    if cell:
        data.setup_margin = sheet.cell(cell.row, cell.column+1).value
    
    # Hold Margin
    cell = find_cell(sheet, "Hold Margin（ns）")
    if not cell:  # 尝试英文括号
        cell = find_cell(sheet, "Hold Margin(ns)")
    if cell:
        data.hold_margin = sheet.cell(cell.row, cell.column+1).value
    
    # Clock Tree Budget -> SM
    ctb_cell = find_cell(sheet, "Clock Tree Budget")
    if ctb_cell:
        # 在CTB列向下搜索SM
        for r in range(ctb_cell.row+1, sheet.max_row+1):
            sm_cell = sheet.cell(r, ctb_cell.column)
            if sm_cell.value == "SM":
                data.ctb.source.min = sheet.cell(r, ctb_cell.column+1).value
                data.ctb.source.max = sheet.cell(r, ctb_cell.column+2).value
                data.ctb.network.min = sheet.cell(r, ctb_cell.column+3).value
                data.ctb.network.max = sheet.cell(r, ctb_cell.column+4).value
                data.ctb.trans.min = sheet.cell(r, ctb_cell.column+5).value
                data.ctb.trans.max = sheet.cell(r, ctb_cell.column+6).value
                data.ctb.skew = sheet.cell(r, ctb_cell.column+7).value
                data.ctb.noise = sheet.cell(r, ctb_cell.column+8).value
                break
    
    return data

def parse_clock_sheet(sheet):
    clock_data = []
    level_cell = find_cell(sheet, "Level")
    
    if level_cell:
        start_row = level_cell.row + 1
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            if not any(row[:10]):  # 跳过空行
                break
            clock_data.append(list(row[:10]))
    
    return clock_data

def main(filename):
    wb = openpyxl.load_workbook(filename)
    
    # 处理pt/sync sheet
    pt_data = None
    sync_data = None
    
    if 'pt' in wb.sheetnames:
        pt_data = parse_pt_sync_sheet(wb['pt'])
    if 'sync' in wb.sheetnames:
        sync_data = parse_pt_sync_sheet(wb['sync'])
    
    # 处理clock sheet
    clock_data = []
    if 'clock' in wb.sheetnames:
        clock_data = parse_clock_sheet(wb['clock'])
    
    # 返回结果 (实际使用时可根据需要处理数据)
    return {
        'pt': pt_data,
        'sync': sync_data,
        'clock': clock_data
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script.py <excel_file>")
        sys.exit(1)
    
    result = main(sys.argv[1])
    
    # 示例输出
    print("PT Data:", result['pt'])
    print("\nSYNC Data:", result['sync'])
    print("\nClock Data (first 5 rows):")
    for row in result['clock'][:5]:
        print(row)