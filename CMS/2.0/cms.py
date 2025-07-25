import sys
import re
import os
import glob
import openpyxl
from dataclasses import dataclass,field

#============================================================
# Define data class
@dataclass
class Path:
    top:str="None"
    top_path:str="None"
    rtl_path:str="None"
    rtl_list:str="None"
    tb_path:str="None"
    cons_path:str="None"
    param:str="None"
    lib_path:str="None"
    lib_target:str="None"
    lib_link:str="None"
    lib_symbol:str="None"
    lib_synthetic:str="None"
@dataclass
class Range:
    min:str="None"
    max:str="None"
@dataclass
class DrivingCell:
    name:str="None"
    input:str="None"
    output:str="None"
    lib:str="None"
@dataclass
class ClockTree:
    source_latency:Range=field(default_factory=Range)
    network_latency:Range=field(default_factory=Range)
    trans:Range=field(default_factory=Range)
    skew:str="None"
    noise:str="None"
@dataclass
class Config:
    path:str="None"
    signal_driving_cell:DrivingCell=field(default_factory=DrivingCell)
    clock_driving_cell:DrivingCell=field(default_factory=DrivingCell)
    signal_trans:Range=field(default_factory=Range)
    output_load:Range=field(default_factory=Range)
    output_trans:Range=field(default_factory=Range)
    output_delay:Range=field(default_factory=Range)
    input_delay:Range=field(default_factory=Range)
    input_trans:Range=field(default_factory=Range)
    setup_margin:str="None"
    hold_margin:str="None"
    sm_ct:ClockTree=field(default_factory=ClockTree)
    md_ct:ClockTree=field(default_factory=ClockTree)
    lg_ct:ClockTree=field(default_factory=ClockTree)
    raw_ct:ClockTree=field(default_factory=ClockTree)

def error_exit(msg):
    print(f"错误：{msg}", file=sys.stderr)
    sys.exit(1)
def standardize_list(list):
    if not list:
        return ""
    return list.replace(","," ").replace(";"," ").replace("\n"," ").replace("\r"," ")
#def is_empty(var):
def parse_path(sheet):
    path=Path()
    path.top = sheet['B1'].value
    if path.top is None or str(path.top).strip() == "":
        error_exit("Not define top module.")
    print(f"Design top module: {path.top}")
    if path.top in os.environ:
        path.top_path = os.environ.get(path.top)
    else:
        path.top_path = sheet['B2'].value
        if path.top_path is None:
            error_exit("Error: Not define design path")

    path.rtl_path = sheet['B3'].value
    path.rtl_path = re.sub(r"\$\{top\}",path.top_path,path.rtl_path,flags=re.IGNORECASE)
    rtl_pattern = ["*.v", "*.sv", "*.vhd", "*.vhdl"]
    path.rtl_list = []
    for pattern in rtl_pattern:
        path.rtl_list.extend(glob.glob(os.path.join(path.rtl_path, pattern)))

    if not path.rtl_list:
        error_exit(f"Not find RTL files in {path.rtl_path} with pattern {rtl_pattern}")
    path.rtl_list = " ".join(os.path.basename(f) for f in path.rtl_list)
    print(f"RTL files found: {path.rtl_list}")

    top_pattern = path.top+"."
    top_match = [item for item in path.rtl_list.split() if top_pattern in item]
    if len(top_match) == 1:
        top_file = top_match[0]
    else:
        print(f"Top-level file: {top_match}")
        error_exit("The top-level file is not unique or not found.")
    print(f"Top-level file: {top_file}")

    param_lines = []
    found_param = False
    with open(path.rtl_path+"/"+top_file,'r',encoding='utf-8') as file:
        for line in file:
            stripped = line.strip()
            if stripped.startswith("//"):
                continue
            if not found_param:
                if "parameter" in stripped:
                    found_param = True
                    param_lines.append(stripped)
            else:
                param_lines.append(stripped)
                if ")" in stripped:
                    break
    path.param = " ".join(param_lines).replace("\n"," ").replace("\r"," ")
    path.param = path.param[len("parameter"):].lstrip()
    path.param = path.param.split(")")[0].rstrip()
    print(f"Parameter of top design: {path.param}")




    path.tb_path = sheet['B4'].value
    path.tb_path = re.sub(r"\$\{top\}",path.top_path,path.tb_path,flags=re.IGNORECASE)

    path.cons_path = sheet['B5'].value
    path.cons_path = re.sub(r"\$\{top\}",path.top_path,path.cons_path,flags=re.IGNORECASE)

    path.lib_path = sheet['B6'].value
    path.lib_path = re.sub(r"\$\{top\}",path.top_path,path.lib_path,flags=re.IGNORECASE)

    path.lib_target = standardize_list(sheet['B9'].value)
    if path.lib_target is None or str(path.lib_target).strip() == "":
        error_exit("Not define target library")
    print(f"Target Library: {path.lib_target}")
    path.lib_link = path.lib_target + " " + standardize_list(sheet['C9'].value) + " *"
    print(f"Link Library: {path.lib_link}")
    path.lib_symbol = standardize_list(sheet['D9'].value)
    if path.lib_symbol is None or str(path.lib_symbol).strip() == "":
        error_exit("Not define symbol library")
    print(f"Symbol Library: {path.lib_symbol}")
    path.lib_synthetic = standardize_list(sheet['E9'].value)
    print(f"Synthetic Library: {path.lib_synthetic}")

    return path

def parse_config(sheet,path):
    config=Config()
    config.path                         = sheet['B1'].value
    config.signal_driving_cell.name     = sheet['B3'].value
    config.signal_driving_cell.output   = sheet['D3'].value
    config.signal_driving_cell.input    = sheet['E3'].value
    config.signal_driving_cell.lib      = sheet['F3'].value
    config.clock_driving_cell.name      = sheet['B4'].value
    config.clock_driving_cell.output    = sheet['D4'].value
    config.clock_driving_cell.input     = sheet['E4'].value
    config.clock_driving_cell.lib       = sheet['F4'].value
    config.signal_trans.min             = sheet['B7'].value
    config.signal_trans.max             = sheet['C7'].value
    config.output_load.min              = sheet['B8'].value
    config.output_load.max              = sheet['C8'].value
    config.output_trans.min             = sheet['B9'].value
    config.output_trans.max             = sheet['C9'].value
    config.output_delay.min             = sheet['B10'].value
    config.output_delay.max             = sheet['C10'].value
    config.input_delay.min              = sheet['B11'].value
    config.input_delay.max              = sheet['C11'].value
    config.input_trans.min              = sheet['B12'].value
    config.input_trans.max              = sheet['C12'].value
    config.setup_margin                 = sheet['B14'].value
    config.hold_margin                  = sheet['B15'].value
    config.sm_ct.source_latency.min     = sheet['B20'].value
    config.sm_ct.source_latency.max     = sheet['C20'].value
    config.sm_ct.network_latency.min    = sheet['D20'].value
    config.sm_ct.network_latency.max    = sheet['E20'].value
    config.sm_ct.trans.min              = sheet['F20'].value
    config.sm_ct.trans.max              = sheet['G20'].value
    config.sm_ct.skew                   = sheet['H20'].value
    config.sm_ct.noise                  = sheet['I20'].value
    config.md_ct.source_latency.min     = sheet['B21'].value
    config.md_ct.source_latency.max     = sheet['C21'].value
    config.md_ct.network_latency.min    = sheet['D21'].value
    config.md_ct.network_latency.max    = sheet['E21'].value
    config.md_ct.trans.min              = sheet['F21'].value
    config.md_ct.trans.max              = sheet['G21'].value
    config.md_ct.skew                   = sheet['H21'].value
    config.md_ct.noise                  = sheet['I21'].value
    config.lg_ct.source_latency.min     = sheet['B22'].value
    config.lg_ct.source_latency.max     = sheet['C22'].value
    config.lg_ct.network_latency.min    = sheet['D22'].value
    config.lg_ct.network_latency.max    = sheet['E22'].value
    config.lg_ct.trans.min              = sheet['F22'].value
    config.lg_ct.trans.max              = sheet['G22'].value
    config.lg_ct.skew                   = sheet['H22'].value
    config.lg_ct.noise                  = sheet['I22'].value
    config.raw_ct.source_latency.min    = sheet['B23'].value
    config.raw_ct.source_latency.max    = sheet['C23'].value
    config.raw_ct.network_latency.min   = sheet['D23'].value
    config.raw_ct.network_latency.max   = sheet['E23'].value
    config.raw_ct.trans.min             = sheet['F23'].value
    config.raw_ct.trans.max             = sheet['G23'].value
    config.raw_ct.skew                  = sheet['H23'].value
    config.raw_ct.noise                 = sheet['I23'].value
    config.path = re.sub(r"\$\{top\}",path.top_path,config.path,flags=re.IGNORECASE)
    return config

def parse_clock(sheet):
    clock_dict  = {}
    row_index   = 4
    columns     = ['level','group','type','period','name','master','jsrc','jmn','jdc','root','comment']
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        row_data = dict(zip(columns,row[:11]))
        name = row_data['name']
        if name:
            clock_dict[name] = row_data
        row_index += 1
    return clock_dict
def parse_rst(sheet):
    rst_dict  = {}
    row_index   = 3
    columns     = ['level','reset','type','edge','clock']
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        row_data = dict(zip(columns,row[:5]))
        reset = row_data['reset']
        if reset:
            rst_dict[reset] = row_data
        row_index += 1
    return rst_dict
def parse_io(sheet):
    io_dict     = {}
    row_index   = 4
    columns     = ['level','pin','direction','clock','trans_min','trans_max','delay_min','delay_max','delay_cmd','load_min','load_max']
    while True:
        row = [cell.value for cell in sheet[row_index]]
        if row[0] is None:
            break
        row_data = dict(zip(columns,row[:11]))
        pin = row_data['pin']
        if pin:
            io_dict[pin] = row_data
        row_index += 1
    return io_dict

def gen_cms_cons_pt(pt_config):
    cons = ""

def gen_cms_cons_synth(synth_config,path):
    cons = "\n#========================================\n#Add LIB"
    cons += f"\nread_db [list {path.lib_target}]"
    cons += f"\nset TOP_MODULE {path.top}"
    cons += f"\nanalyze -f sverilog [list {path.rtl_list}]"
    cons += f"\nelaborate $TOP_MODULE -parameter \"{path.param}\""
    cons += f"\ncurrent_design TOP_MODULE"
    cons += "\nif {[check_design] == 0} {\n   echo \"Check Design Error!\";\n   exit;\n}"
    cons += "\nreset_design"
    cons += "\nuniquify"       
    cons += "\nset uniquify_naming_style   \“%s_%d\”"
    cons += "\nwrite -f ddc -hierarchy -output ${UNMAPPED_PATH}/${TOP_MODULE}.ddc"
    cons += f"\nsource {path.cons_path}/{path.top}.tcl"
    cons += "\ncheck_timing\nset_host_option -max_cores 8"
    cons += "\ncompile -map_effort high"
    #cons += f"\n{synth_config.path}"

    
    print(cons)

def gen_cms_cons_clk(clock_dict):
    cons = "\n#========================================\n#Clock Constraint"
    if not clock_dict:
        print("No clock data found.")
        return
    
    print(f"Starting to generate CMS clock constraints for {len(clock_dict)} clocks.")
    print("-"*20)

    for name,row_data in clock_dict.items():
        
        if clock_dict[name]['root'] is None:
            cons += f"\ncreate_clock -name {name} -period {clock_dict[name]['period']}"
        else:
            if "/" in clock_dict[name]['root']:
                pin_or_port = "pins"
            else:
                pin_or_port = "ports"
            
            if clock_dict[name]['master'] is None:
                cons += f"\ncreate_clock -name {name} -period {clock_dict[name]['period']} [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_ideal_network [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_dont_touch_network [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_drive [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_clock_uncertaity  -setup $CLK_SKEW [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_clock_transition  -max $CLK_TRAN [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_clock_latency -source -max $CLK_SRC_LATENCY [get_{pin_or_port} {clock_dict[name]['root']}]"
                cons += f"\nset_clock_latency -max $CLK_LATENCY [get_{pin_or_port} {clock_dict[name]['root']}]"
            else:
                #mst_clk = 
                mst_source = clock_dict[clock_dict[name]['master']]['root']
                if not "/" in mst_source:
                    mst_source = f"[get_port {mst_source}]"
                if match := re.match(r"-div\s+(\d+)$",clock_dict[name]['period']):
                    generated_period = f"-divide_by {match.group(1)}"
                elif match := re.match(r"-multi\s+(\d+)$",clock_dict[name]['period']):
                    generated_period = f"-multiply_by {match.group(1)}"
                elif match := re.match(r"-edges\s+\{\s*\d+\s+\d+\s+\d+\s*\}$",clock_dict[name]['period']):
                    generated_period = clock_dict[name]['period']
                else:
                    generated_period = clock_dict[name]['period']
                cons += f"\ncreate_generated_clock -name {name} {generated_period} -source {mst_source} [get_{pin_or_port} {clock_dict[name]['root']}]" 
    
    return cons


def gen_cms_cons_rst(rst_dict,clock_dict):
    cons = "\n#========================================\n#Reset Constraint"
    if not rst_dict:
        print("No Reset data found.")
        return
    
    print(f"Starting to generate CMS Reset constraints for {len(rst_dict)} reset.")
    print("-"*20)
    
    for reset,row_data in rst_dict.items():
        cons += f"\nset_ideal_network [get_port {reset}]"
        cons += f"\nset_dont_touch_network [get_port {reset}]"
        cons += f"\nset_drive 0 [get_port {reset}]"
    
    return cons

def gen_cms_cons_io(io_dict,clock_dict):
    cons = "\n#========================================\n#IO Constraint"
    if not io_dict:
        print("No IO data found.")
        return
    
    print(f"Starting to generate CMS IO constraints for {len(io_dict)} pins.")
    print("-"*20)
    #for row in enumerate(clock_list,start=1):
    for pin,row_data in io_dict.items():
        if io_dict[pin]['direction'] in ("input","in"):
            if io_dict[pin]['delay_min'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_min']
                cons += f"\nset_input_delay -clock {io_dict[pin]['clock']} -min {delay} [get_ports {pin}]"
            if io_dict[pin]['delay_max'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_max']
                cons += f"\nset_input_delay -clock {io_dict[pin]['clock']} -max {delay} [get_ports {pin}]"
        elif io_dict[pin]['direction'] in ("output","out"):
            if io_dict[pin]['delay_min'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_min']
                cons += f"\nset_output_delay -clock {io_dict[pin]['clock']} -min {delay} [get_ports {pin}]"
            if io_dict[pin]['delay_max'] is not None:
                delay = clock_dict[io_dict[pin]['clock']]['period'] * io_dict[pin]['delay_max']
                cons += f"\nset_output_delay -clock {io_dict[pin]['clock']} -max {delay} [get_ports {pin}]"
            if io_dict[pin]['load_min'] is not None:
                load = io_dict[pin]['load_min']
                cons += f"\nset_load -min {load} [get_ports {pin}]"
            if io_dict[pin]['load_max'] is not None:
                load = io_dict[pin]['load_max']
                cons += f"\nset_load -max {load} [get_ports {pin}]"
    return cons

def main(filename):
    cons = ""
    wb = openpyxl.load_workbook(filename)

    if 'path' in wb.sheetnames:
        path = parse_path(wb['path'])
    if 'pt' in wb.sheetnames:
        pt_config = parse_config(wb['pt'],path)
        #cons_pt = gen_cms_cons_pt(pt_config,design)
    if 'synth' in wb.sheetnames:
        synth_config = parse_config(wb['synth'],path)
        cons_synth = gen_cms_cons_synth(synth_config,path)

    if 'clock' in wb.sheetnames:
        clock_dict = parse_clock(wb['clock'])
        #print (clock_list)
        cons_clk = gen_cms_cons_clk(clock_dict)
    if 'reset' in wb.sheetnames:
        rst_dict = parse_rst(wb['reset'])
        cons_rst = gen_cms_cons_rst(rst_dict,clock_dict)
    if 'io' in wb.sheetnames:
        io_dict = parse_io(wb['io'])
        #print (io_list)
        cons_io = gen_cms_cons_io(io_dict,clock_dict)
    #print(pt_config)
    #print(synth_config)
    #print(clock_list)
    

    cons += cons_clk
    cons += cons_rst
    cons += cons_io
    #print (cons)

    
    wb.close()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python cms.py <excel_file>")
        sys.exit(1)
    
    table_path = sys.argv[1]
    try:
        result = main(table_path)
    except Exception as e:
        print("Error:", e)